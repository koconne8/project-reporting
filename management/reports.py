import datetime
from django.db import connection
from django.shortcuts import HttpResponse
from openpyxl import Workbook
from holidays import GetHolidays
from openpyxl.styles import Color, Font, PatternFill
from openpyxl.styles import colors
from django.contrib.auth.decorators import login_required

@login_required
def WeeklyReportFromUrl(request):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="WeeklyReport.xls"'

    return GenerateWeeklyReport(response)

def GenerateWeeklyReport(file_out):
    # get a list of all projects we should include in the report
    cur = connection.cursor()

    cur.execute("select sum(hours), projects.name, project_id "
                "FROM time_entries "
                "INNER JOIN projects ON time_entries.project_id = projects.id "
                "INNER JOIN custom_values ON custom_values.customized_id = time_entries.project_id "
                "WHERE custom_values.custom_field_id = 19 "
                "AND custom_values.value = '1' "
                "AND spent_on >= (current_date - interval '7 days') "
                "GROUP BY projects.name, project_id;")

    projects = cur.fetchall()

    project_list = []
    for project in projects:
        proj = {
            'name': project[1],
            'id': project[2],
            'hours': project[0]
        }

        # get the budget
        cur.execute("SELECT value FROM custom_values WHERE custom_field_id = 12 AND customized_id = %(project_id)s;" % {'project_id': project[2]})
        proj['budget'] = (float(cur.fetchone()[0]))

        # get the accumulated amount spent
        cur.execute("SELECT value FROM custom_values WHERE custom_field_id = 13 AND customized_id = %(project_id)s;" % {
            'project_id': project[2]})
        proj['spent'] = (float(cur.fetchone()[0]))

        # get the start date
        cur.execute("SELECT value FROM custom_values WHERE custom_field_id = 15 AND customized_id = %(project_id)s;" % {
            'project_id': project[2]})
        proj['start_date'] = cur.fetchone()[0]

        # get the end date
        cur.execute("SELECT value FROM custom_values WHERE custom_field_id = 16 AND customized_id = %(project_id)s;" % {
            'project_id': project[2]})
        proj['end_date'] = cur.fetchone()[0]

        # get the FTE effort for TODAY
        cur.execute("SELECT SUM(percentage) FROM project_distribution WHERE \"from\" <= now() AND \"to\" >= now() AND project = %(proj)s;" % {'proj': proj['id']})
        effort = cur.fetchone()
        if len(effort) > 0:
            proj['fte'] = effort[0]
        else:
            proj['fte'] = '0'

        # run through each day between now and the project's end date.
        current_day = datetime.datetime.now()
        end_date = datetime.datetime.strptime(proj['end_date'], '%Y-%m-%d')

        # remember projected spending is IN ADDITION to what's already spent...
        projected_spending = proj['spent']
        while current_day <= end_date:
            # is this a working day?
            # ...first check for holidays
            if current_day in GetHolidays(current_day.year):
                current_day = current_day + datetime.timedelta(days=1)
                continue

            # ...then check if we are on a weekend
            if current_day.weekday() > 4:
                current_day = current_day + datetime.timedelta(days=1)
                continue


            # for each day, figure out the FTE effort
            cur.execute(
                "SELECT SUM(percentage) FROM project_distribution WHERE \"from\" <= '%(day)s' AND \"to\" >= '%(day)s' AND project = %(proj)s;" % {
                    'proj': proj['id'], 'day': current_day})
            effort = cur.fetchone()
            if len(effort) > 0:
                effort = effort[0]
            else:
                effort = 0

            # go get the charge rate for this day
            cur.execute(
                "SELECT rate FROM charge_rates WHERE start_date <= '%(date)s' AND end_date >= '%(date)s' AND category = 'Programming' AND internal = TRUE;" % {
                    'date': current_day
                }
            )
            rate = cur.fetchone()
            if len(rate) > 0:
                rate = rate[0]
            else:
                rate = 0

            # now we can add the product of today's rate and effort to the projected spending
            projected_spending += (float(rate) * float(effort))

            current_day = current_day + datetime.timedelta(days=1)

        proj['projected_spending'] = projected_spending

        # the ratio is: [projected spending] / [budget]
        proj['projected_ratio'] = '{:,.2f}%'.format((projected_spending / float(proj['budget']) * 100))

        project_list.append(proj)


    # create an active worksheet for our excel file
    wb = Workbook()
    ws = wb.active

    # The Headers
    ws['B1'] = 'Project'
    ws['C1'] = 'Budget'
    ws['D1'] = 'Spent'
    ws['E1'] = 'FTE Effort'
    ws['F1'] = 'End Date'
    ws['G1'] = 'Projected Spending'
    ws['H1'] = 'Projected Spending Ratio'
    ws['I1'] = 'Billed Hours (past 7 days)'

    # style headers
    header_font = Font(color=colors.WHITE)
    header_bg = PatternFill(fill_type=None, start_color='000000', end_color='000000')
    # font color
    # ws['B1'].font.color.index = 'FFFFFF'
    # ws['C1'].font.color.index = 'FFFFFF'
    # ws['D1'].font.color.index = 'FFFFFF'
    # ws['E1'].font.color.index = 'FFFFFF'
    # ws['F1'].font.color.index = 'FFFFFF'
    # ws['G1'].font.color.index = 'FFFFFF'
    # ws['H1'].font.color.index = 'FFFFFF'
    # ws['I1'].font.color.index = 'FFFFFF'

    # background fill type
    # ws['B1'].style.fill.fill_type = Fill.FILL_SOLID
    # ws['C1'].style.fill.fill_type = Fill.FILL_SOLID
    # ws['D1'].style.fill.fill_type = Fill.FILL_SOLID
    # ws['E1'].style.fill.fill_type = Fill.FILL_SOLID
    # ws['F1'].style.fill.fill_type = Fill.FILL_SOLID
    # ws['G1'].style.fill.fill_type = Fill.FILL_SOLID
    # ws['H1'].style.fill.fill_type = Fill.FILL_SOLID
    # ws['I1'].style.fill.fill_type = Fill.FILL_SOLID

    # background fill color
    # ws['B1'].style.fill.start_color.index = '000000'
    # ws['C1'].style.fill.start_color.index = '000000'
    # ws['D1'].style.fill.start_color.index = '000000'
    # ws['E1'].style.fill.start_color.index = '000000'
    # ws['F1'].style.fill.start_color.index = '000000'
    # ws['G1'].style.fill.start_color.index = '000000'
    # ws['H1'].style.fill.start_color.index = '000000'
    # ws['I1'].style.fill.start_color.index = '000000'


    working_row = 2

    print project_list

    for project in project_list:
        ws['B'+str(working_row)] = project['name']
        ws['C'+str(working_row)] = '${:,.2f}'.format(project['budget'])
        ws['D'+str(working_row)] = '${:,.2f}'.format(project['spent'])
        ws['E'+str(working_row)] = project['fte']
        ws['F'+str(working_row)] = project['end_date']
        ws['G'+str(working_row)] = '${:,.2f}'.format(project['projected_spending'])
        ws['H'+str(working_row)] = project['projected_ratio']
        ws['I'+str(working_row)] = project['hours']


        working_row += 1

    wb.save(file_out)
    return file_out